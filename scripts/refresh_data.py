"""
refresh_data.py

Self-serve pipeline: parses Database.xlsx directly and regenerates
data/mounts.json, data/artifacts.json, data/relics.json, plus extracts
all embedded images into assets/images/{mounts,artifacts}/.

This replaces needing to hand the spreadsheet to Claude for routine
updates — run this whenever you've edited the sheet, then run build.py,
then commit + push as usual.

Usage:
    python3 scripts/refresh_data.py path/to/Database.xlsx

Expected sheets (case-sensitive):
    "Mount Collection"          — Name, Rarity, Image (embedded pictures)
    "Mounts StarUpAwakening"    — per-mount star/awaken stat + effect grid
    "Artifact Collection"       — Name, Rarity, Image (embedded pictures)
    "Artifacts StarUpAwakening" — per-artifact star/awaken stat + effect grid
    "Relics"                    — Relic name, Tier, Image, 0-10 star columns

Safe to re-run any time — it's a full regeneration from the sheet, not an
incremental patch. Existing relics.json fields not present in the sheet
(type, effect, pve, pvp, id) are preserved by matching on relic name;
everything else (mounts.json, artifacts.json) is fully replaced since the
sheet is the single source of truth for those.
"""

import sys
import re
import json
import io
from pathlib import Path

try:
    import openpyxl
    from PIL import Image
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl pillow --break-system-packages")
    sys.exit(1)

PROJECT_DIR = Path(__file__).parent.parent
DATA_DIR = PROJECT_DIR / "data"
ASSETS_DIR = PROJECT_DIR / "assets" / "images"

STAR_LABELS = ['1', '2', '3', '4', '5']
AWAKEN_LABELS = [f'A{i}' for i in range(1, 11)]
MOUNT_RARITIES = {'Uncommon', 'Rare', 'Epic', 'Legendary', 'Mythic', 'Immortal', 'Transcendent'}
ARTIFACT_RARITIES = {'Legendary', 'Mythic', 'Immortal', 'Transcendent'}

STAT_NAME_MAP = {
    'hp%': 'hp_pct', 'hp': 'hp', 'atk%': 'atk_pct', 'atk': 'atk',
    'def%': 'def_pct', 'def': 'def', 'tenacity': 'tenacity',
    'armor break': 'armor_break', 'counter rate': 'counter_rate',
    'combo rate': 'combo_rate', 'crit rate': 'crit_rate', 'crit dmg': 'crit_dmg',
    'penetration': 'penetration', 'block': 'block', 'suppression': 'suppression',
}


def slugify(name):
    s = (name or '').lower()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return s.strip('-')


def norm_name(n):
    """Loose name match to catch spacing/punctuation drift between sheet
    edits and previously tracked names (e.g. 'Flyer No. 1' vs 'Flyer No.1')."""
    return re.sub(r'[^a-z0-9]', '', (n or '').lower())


def clean_text(text):
    if text is None or text == '—':
        return None
    return str(text).strip()


def parse_stat_block(text):
    if text is None or text == '—':
        return {}
    out = {}
    for line in str(text).split('\n'):
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^(.+?)\s*([+\-])\s*([\d.]+)\s*(%?)$', line)
        if not m:
            out.setdefault('_unparsed', []).append(line)
            continue
        name, sign, num, pct = m.groups()
        key = re.sub(r'[^a-z0-9]+', '_', name.strip().lower()).strip('_')
        val = float(num) if '.' in num else int(num)
        if sign == '-':
            val = -val
        out[key + ('_pct' if pct else '')] = val
    return out


def parse_relic_stat_cell(text):
    """Relics sheet uses 'Stat=Value\\n Stat2=Value2' format, distinct from
    the +/- delta format used on mount/artifact sheets."""
    if text is None:
        return {}
    out = {}
    for line in str(text).split('\n'):
        line = line.strip()
        if not line or '=' not in line:
            continue
        name, val = line.split('=', 1)
        key = STAT_NAME_MAP.get(name.strip().lower())
        if not key:
            key = re.sub(r'[^a-z0-9]+', '_', name.strip().lower()).strip('_')
        try:
            out[key] = float(val) if '.' in val else int(val)
        except ValueError:
            out[key] = val.strip()
    return out


def extract_collection_sheet(wb, sheet_name, out_dir):
    """Parses a *_Collection sheet (Name, Rarity, Image) and extracts
    embedded images. Returns list of {n, tier} in row order, and flags
    any row whose image anchor collides with another row (likely a
    misplaced paste in the source sheet — needs manual review)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    items = []
    for row in rows[1:]:
        if row[0]:
            items.append({'n': row[0], 'tier': row[1]})

    out_dir.mkdir(parents=True, exist_ok=True)
    anchor_rows = {}
    warnings = []
    for img in ws._images:
        frm = img.anchor._from
        sheet_row = frm.row + 1
        if sheet_row in anchor_rows:
            warnings.append(
                f"  [!] Two images both anchored at row {sheet_row} in '{sheet_name}' — "
                f"one is likely misplaced. Check {rows[sheet_row-1][0] if sheet_row-1 < len(rows) else '?'} "
                f"and the row below/above it manually."
            )
        anchor_rows.setdefault(sheet_row, []).append(img)

    for sheet_row, imgs in anchor_rows.items():
        if sheet_row - 1 >= len(rows) or not rows[sheet_row - 1][0]:
            warnings.append(f"  [!] Image anchored at row {sheet_row} in '{sheet_name}' has no matching name — skipped.")
            continue
        name = rows[sheet_row - 1][0]
        slug = slugify(name)
        img = imgs[0]  # if collided, still save the first; the warning above flags it for manual check
        data = img.ref.getvalue() if hasattr(img.ref, 'getvalue') else img._data()
        pil_img = Image.open(io.BytesIO(data)).convert('RGBA').resize((96, 96), Image.LANCZOS)
        pil_img.save(out_dir / f'{slug}.webp', 'WEBP', quality=90)

    return items, warnings


def parse_starupawaken_sheet(wb, sheet_name, rarities):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    out = {}
    i = 0
    while i < len(rows):
        label = rows[i][0]
        if label in rarities or label is None or label == 'Star Up Stats':
            i += 1
            continue
        if label in ('Delta from Base', 'Effect'):
            i += 1
            continue

        name_row = rows[i]
        delta_row = rows[i + 1] if i + 1 < len(rows) and rows[i + 1][0] == 'Delta from Base' else None
        effect_row = rows[i + 2] if i + 2 < len(rows) and rows[i + 2][0] == 'Effect' else None
        name = name_row[0]

        star_base = parse_stat_block(name_row[1])
        star_deltas = {STAR_LABELS[k]: parse_stat_block(delta_row[2 + k]) for k in range(5)} if delta_row else {}
        awaken_base = parse_stat_block(name_row[9])
        awaken_deltas = {AWAKEN_LABELS[k]: parse_stat_block(delta_row[10 + k]) for k in range(10)} if delta_row else {}

        base_effect = None
        awaken_base_effect = None
        star_effects = {}
        awaken_effects = {}
        if effect_row:
            base_effect = clean_text(effect_row[1])
            awaken_base_effect = clean_text(effect_row[9])
            star_effects['0'] = base_effect
            for k, lvl in enumerate(STAR_LABELS):
                e = clean_text(effect_row[2 + k])
                star_effects[lvl] = e if e else base_effect
            for k, lvl in enumerate(AWAKEN_LABELS):
                e = clean_text(effect_row[10 + k])
                if e:
                    awaken_effects[lvl] = e

        out[name] = {
            'base_effect': base_effect,
            'awaken_base_effect': awaken_base_effect,
            'star_up': {'base_stats': star_base, 'deltas': star_deltas},
            'awaken': {'base_stats': awaken_base, 'deltas': awaken_deltas},
            'star_effects': star_effects,
            'awaken_effects': awaken_effects,
        }
        i += 3
    return out


RELIC_NAME_ALIASES = {
    # "Relic Equip effect" sheet naming vs. canonical Relics sheet naming
    # (typos / naming drift between the two sheets)
    'broze statue of shiva as nataraja': 'Bronze Statue of Shiva as Nataraja',
}


def extract_relic_images(wb, sheet_name, out_dir):
    """Same anchor-row extraction approach as Collection sheets, but for the
    Relics sheet's own Image column (col C)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    out_dir.mkdir(parents=True, exist_ok=True)
    anchor_rows = {}
    warnings = []
    for img in ws._images:
        frm = img.anchor._from
        sheet_row = frm.row + 1
        anchor_rows.setdefault(sheet_row, []).append(img)

    saved = 0
    for sheet_row, imgs in anchor_rows.items():
        if len(imgs) > 1:
            name_here = rows[sheet_row - 1][0] if sheet_row - 1 < len(rows) else '?'
            warnings.append(f"  [!] Two images anchored at row {sheet_row} in '{sheet_name}' — check {name_here} and neighboring rows manually.")
        if sheet_row - 1 >= len(rows) or not rows[sheet_row - 1][0]:
            continue
        name = rows[sheet_row - 1][0]
        slug = slugify(name)
        data = imgs[0].ref.getvalue() if hasattr(imgs[0].ref, 'getvalue') else imgs[0]._data()
        pil_img = Image.open(io.BytesIO(data)).convert('RGBA').resize((96, 96), Image.LANCZOS)
        pil_img.save(out_dir / f'{slug}.webp', 'WEBP', quality=90)
        saved += 1
    return saved, warnings


def parse_relic_equip_effect_sheet(wb, existing_names):
    """Parses the optional 'Relic Equip effect' sheet: three side-by-side
    relic blocks per row-group (cols A-D, F-I, K-N), each with its own
    interspersed rarity headers. Returns (matched_dict, unmatched_names)."""
    if 'Relic Equip effect' not in wb.sheetnames:
        return {}, []

    ws = wb['Relic Equip effect']
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # skip the 2 header rows
    RARITY_TOKENS = {'Rare', 'Epic', 'Legendary', 'Mythic'}

    def parse_block(col_start):
        out = {}
        for row in rows:
            name = row[col_start]
            if not name or name in RARITY_TOKENS:
                continue
            out[name.strip()] = {
                'effect_base': clean_text(row[col_start + 1]),
                'effect_5star': clean_text(row[col_start + 2]),
                'effect': clean_text(row[col_start + 3]),  # 10★, matches existing field's meaning
            }
        return out

    raw = {}
    for start in [0, 5, 10]:
        raw.update(parse_block(start))

    matched = {}
    unmatched = []
    for name, effects in raw.items():
        if name in existing_names:
            matched[name] = effects
            continue
        alias = RELIC_NAME_ALIASES.get(name.lower())
        if alias and alias in existing_names:
            matched[alias] = effects
            continue
        unmatched.append(name)

    return matched, unmatched


def build_mounts_or_artifacts(kind, collection_sheet, starup_sheet, rarities):
    import openpyxl as ox
    wb = ox.load_workbook(sys.argv[1], data_only=True)

    print(f"\n--- {kind} ---")
    items, img_warnings = extract_collection_sheet(wb, collection_sheet, ASSETS_DIR / kind)
    stats_by_name = parse_starupawaken_sheet(wb, starup_sheet, rarities)

    stats_by_norm = {norm_name(n): v for n, v in stats_by_name.items()}

    out = [{'idx': 0, 'n': 'None', 'tier': '', 'base_effect': None}]
    unmatched = []
    for idx, item in enumerate(items, start=1):
        stats = stats_by_norm.get(norm_name(item['n']))
        entry = {'idx': idx, 'n': item['n'], 'tier': item['tier']}
        if stats:
            entry.update(stats)
        else:
            unmatched.append(item['n'])
        out.append(entry)

    print(f"  {len(items)} items in Collection sheet, {len(stats_by_name)} in StarUpAwakening sheet")
    if unmatched:
        print(f"  [!] {len(unmatched)} item(s) have no matching StarUpAwakening data: {unmatched}")
    for w in img_warnings:
        print(w)

    json.dump({kind: out}, open(DATA_DIR / f'{kind}.json', 'w'), indent=2, ensure_ascii=False)
    print(f"  wrote data/{kind}.json ({len(out)} entries)")


def build_relics():
    import openpyxl as ox
    wb = ox.load_workbook(sys.argv[1], data_only=True)
    ws = wb['Relics']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = rows[0]
    star_cols = list(range(3, 14))  # 0★..10★

    print("\n--- Relics ---")
    existing_path = DATA_DIR / 'relics.json'
    existing = json.load(open(existing_path)) if existing_path.exists() else {'relics': [], 'relic_sets': {}}
    existing_by_name = {r['n']: r for r in existing['relics']}

    img_saved, img_warnings = extract_relic_images(wb, 'Relics', ASSETS_DIR / 'relics')
    print(f"  extracted {img_saved} relic image(s)")
    for w in img_warnings:
        print(w)

    effect_updates, effect_unmatched = parse_relic_equip_effect_sheet(wb, set(existing_by_name.keys()) | {row[0] for row in rows[1:] if row[0]})
    if effect_updates:
        print(f"  merged richer effect text (base/5★/10★) for {len(effect_updates)} relic(s) from 'Relic Equip effect' sheet")
    if effect_unmatched:
        print(f"  [!] {len(effect_unmatched)} relic(s) in 'Relic Equip effect' sheet not found in the Relics sheet at all "
              f"(likely new relics missing star-progression data — add manually or ask Claude): {effect_unmatched}")

    new_relics = []
    unmatched_existing = set(existing_by_name.keys())
    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        tier = row[1]
        per_star = [parse_relic_stat_cell(row[c]) for c in star_cols]
        # transpose: stat -> [val at 0★, 1★, ..., 10★]
        all_stats = set()
        for s in per_star:
            all_stats.update(s.keys())
        star_stats = {stat: [s.get(stat, 0) for s in per_star] for stat in all_stats}

        prior = existing_by_name.get(name, {})
        unmatched_existing.discard(name)
        entry = {
            'id': prior.get('id') or None,  # None (not '') when unknown — empty strings collide as a shared app state key
            'n': name,
            'type': prior.get('type', ''),
            'effect': prior.get('effect', ''),
            'pve': prior.get('pve', ''),
            'pvp': prior.get('pvp', ''),
            'rarity': tier,
            'star_stats': star_stats,
        }
        if name in effect_updates:
            eff = effect_updates[name]
            entry['effect'] = eff['effect'] or entry['effect']
            entry['effect_base'] = eff['effect_base']
            entry['effect_5star'] = eff['effect_5star']
        new_relics.append(entry)

    print(f"  {len(new_relics)} relics in sheet")
    if unmatched_existing:
        print(f"  [!] {len(unmatched_existing)} previously-tracked relic(s) not found in this sheet "
              f"(kept out of the new file — check for renames): {sorted(unmatched_existing)}")

    json.dump({'relics': new_relics, 'relic_sets': existing.get('relic_sets', {})},
               open(existing_path, 'w'), indent=2, ensure_ascii=False)
    print(f"  wrote data/relics.json ({len(new_relics)} entries, relic_sets preserved as-is)")


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 scripts/refresh_data.py path/to/Database.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Refreshing from {xlsx_path}")
    build_mounts_or_artifacts('mounts', 'Mount Collection', 'Mounts StarUpAwakening', MOUNT_RARITIES)
    build_mounts_or_artifacts('artifacts', 'Artifact Collection', 'Artifacts StarUpAwakening', ARTIFACT_RARITIES)
    build_relics()

    print("\nDone. Next steps:")
    print("  python3 build.py")
    print("  git add .")
    print("  git commit -m \"update database\"")
    print("  git push")


if __name__ == '__main__':
    main()
